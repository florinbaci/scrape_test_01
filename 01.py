from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import re
import ast
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# import sqlite3

# import csv
# import pandas as pd
# from pandas import DataFrame

# Setting up the environment and getting the web page raddy to scrap
driver = webdriver.Chrome('C:/Users/User/AppData/Local/Programs/Python/Python38/Scripts/chromedriver.exe')
driver.maximize_window()
url = "https://www.betfair.ro/exchange/plus/ro/curse-de-cai-pariuri-7/next"
driver.get(url)

# print(driver.page_source)

# while True:

# Get the list of events and some other tabs from the page
race_list = []
races_class = WebDriverWait(driver, 30).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "node")))
for race in races_class:
    race_list.append(race.text)
#     print(race.text)
# print(race_list)

# Because the list is in the second element it had to be separated and then use re library to select each event
# the re library is to select only te time "tag", the rest of the elements don't have the hour pattern
# THINGS TO SOLVE IN HERE:
# SOMETIMES THE PAGE DOESN'T LOAD PROPERLY AND FOR THIS REASON CAN'T FIND THE SECOND ELEMENT
# THERE IS TRY AND EXCEPT BUT DOESN'T WORK PROPERLY, WHEN CAN'T FIND THE SECOND ELEMENT
# THE PROBLEM GOES DOWN TO THE NEXT STATEMENT AND GIVES AN ERROR BECAUSE IT CAN'T FIND THE EVENTS
# VARIABLE
try:
    race_list = race_list[2].split()
    scheduled_hours = [i for i in race_list if re.search(r'\d\d:\d\d', i)]
    events = scheduled_hours[:7]  # only 10 events
    # print(events)
except:
    "IndexError: list index out of range"

# Preparing each event for the scrap: obtaining the sum that it's been bet on the event, the title and the buttons
# obtaining all the links from the horse racing page
# PROBLEM TO BE DEALT WITH:
# WHEN IT LOOPS THROUGH ALL THE LINKS, IT TAKES THOSE RACES WITH THE SAME NAMES THAT ARE A DAY APART
# FOR NOW THIS ISN'T A PROBLEM BECAUSE THE RACES THAT ARE ON THE NEXT DAY DON'T MEET THE REQUIREMENT
# OF HAVING THE TOTAL SUM OF MONEY OVER 5000 LEI, BUT JUST A SMALL BUG TO HAVE IN MIND
race_title = []
race_link = []

for event in events:
    races = driver.find_elements_by_partial_link_text(event)
    for race in races:
        race_title.append(race.text)
        race_link.append(race.get_attribute('href'))

# print(race_title)
# for link in race_link:
#     print(link)
# print(race_link)

# Find the race winner
# HERE IS A DELAY IMPLEMENTED, AFTER EACH ROW CHECKED WAITS 5 SEC FOR THE JOKEY TO BE
# IDENTIFIED, SO SAY IF THE WINNER IS AT POSITION 5, IT WILL TAKE 20 TO 25 SEC GET
# THE CORRECT ANSWER

def race_finished():

    # Make a list of last 5 links from excel to open them and check the winner
    # Load the Excel file and make it active
    workbook_name = 'data.xlsx'
    wb = load_workbook(workbook_name)
    ws = wb.active

    # Because openpyxl doesn't have a module that let you select last n rows
    # I have to get the length of the C column for which I subtract the rows
    # that have to be checked, in this case I got the last 7 rows
    col_c = ws['c']
    len_table = int(len(col_c))
    # print print(len_table - 5)
    fifth_row = int(len_table - 5)

    # Select the columns and rows to be checked
    row = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=8, max_col=8)

    # Create an empty list to store the links that has to be checked
    links_to_check = []

    for link in row:
        links_to_check.append(link.value)

    print(links_to_check)

    # wb.save(filename=workbook_name)

    # jockey_names = WebDriverWait(driver, 30).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "name")))
    # jockey_names_list = []
    # for jockey in jockey_names:
    #     jokey_01 = jockey.text
    #     jockey_names_list.append(jokey_01)
    #     race_winner = WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.CLASS_NAME, "runner-winner")))
    #     race_win01: str = '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[' \
    #                       '1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[' \
    #                       '2]/div/div/div/table/tbody/tr['
    #     race_win02: str = ']/td/div[1]/div[2]/div'
    # # print(jockey_names_list)
    # position_final = []
    #
    # for position in range(len(jockey_names)):
    #     race_w = race_win01 + str(position + 1) + race_win02
    #     try:
    #         winner = WebDriverWait(driver, 5).until(ec.presence_of_element_located((By.XPATH, race_w)))
    #         if winner.text == "Câştigător":
    #             position_final.append(position)
    #         break
    #     except:
    #         "TimeoutException"
    #
    # venue_name = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "venue-name")))
    # venue_name_final = venue_name.text
    # venue_name_final_race_fin01 = str(venue_name_final.split()[1:-1])
    # venue_name_final_race_fin02 = (' '.join(ast.literal_eval(venue_name_final_race_fin01)))
    #
    # winner_final = jockey_names_list[int(position_final[0])]
    # # print(venue_name_final)
    # # print(winner_final)
    #
    # # venue_date_race_fin = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "event-date")))
    # # venue_date_final_race_fin = venue_date_race_fin.text
    # # venue_date_final_race_fin01 = venue_date_final_race_fin[3:]
    # # print(venue_date_final01)
    #
    # venue_time_race_fin = venue_name_final.split()[0]
    # venue_time_final_race_fin = datetime.datetime.strptime(venue_time_race_fin, '%H:%M').time()
    # # print(venue_time_final_race_fin)
    #
    # race_final_data = [venue_name_final_race_fin02, venue_time_final_race_fin]
    # print(race_final_data)
    # print(winner_final)
    # print("in")
    #
    # workbook_name = 'data.xlsx'
    # wb = load_workbook(workbook_name)
    # ws = wb.active
    #
    # col_c = ws['c']
    # len_table = int(len(col_c))
    # # print print(len_table - 5)
    # fifth_row = int(len_table - 7)
    #
    # # row_range = ws[fifth_row: len_table]
    # # print(row_range)
    #
    # col_b = int(fifth_row)
    # col_c = int(len_table)
    # # col_bf = str('B') + str(col_b)
    # # col_cf = str('C') + str(col_c)
    # # # print(col_bf)
    # # # print(col_cf)
    #
    # time_race_list = []
    #
    # row_time_race = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=2, max_col=3)
    #
    # for time, race in row_time_race:
    #     time_race_list.append([time.value, race.value])
    #
    # jokey_list = []
    #
    # row_jokey = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=7, max_col=7)
    #
    # for jokey in row_jokey:
    #     for jokey_value in jokey:
    #         jokey_list.append(jokey_value.value)
    #
    # for time_race_jokey in range(len(time_race_list)):
    #     time_race_list[time_race_jokey].append(jokey_list[time_race_jokey])

    # print(time_race_list)

    # print(jokey_list)

    # def index_in_list_of_lists(time_race_list, winner_final):
    #     for i, lst in enumerate(time_race_list):
    #         if winner_final in lst:
    #             print("1")
    #             break
    #         else:
    #             print("0")
    #             # raise ValueError("%s not in list_of_lists" % winner_final)
    #
    #     return i, lst.index(winner_final)
    #
    # # print(index_in_list_of_lists(time_race_list, winner_final))
    #
    # h = index_in_list_of_lists(time_race_list, winner_final)
    # print(h[0])


# This function takes the races that meet the requirements
# IN HERE SHOULD BE THE "PRESSING OF THE BUTTONS ACTION"
# AFTER ALL CONDITIONS ARE CHECKED TO PLACE THE BET AND CALCULATE IF IT DOES HAVE TO
# INCREASE THE MONEY THAT HAVE TO BE PLACED ON A BET OR CO REVERSE TO THE "BASE"
# FOR NOW THIS COME SECOND ORDER OF IMPORTANCE DUE TO THE FACT THE WE WANT TO TEST
# THE HYPOTHESIS FIRST

def race_to_start():

    jockey_names = WebDriverWait(driver, 30).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "name")))
    jockey_names_list = []
    for jockey in jockey_names:
        jokey_01 = jockey.text
        jockey_names_list.append(jokey_01)

    combined = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "total-matched")))
    event_money = combined.text
    event_money_int = int(event_money.split()[1].replace(',', ''))

    if event_money_int > 5000:

        race_to_check = race

        venue_name = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "venue-name")))
        venue_name_final = venue_name.text

        venue_date = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "event-date")))
        venue_date_final = venue_date.text
        venue_date_final01 = venue_date_final[3:]
        # print(venue_date_final01)

        venue_time = venue_name_final.split()[0]
        venue_time_final = datetime.datetime.strptime(venue_time, '%H:%M').time()
        # print(venue_time_final)

        # print(type(venue_time))

        venue_name_final01 = str(venue_name_final.split()[1:-1])
        venue_name_final02 = (' '.join(ast.literal_eval(venue_name_final01)))
        # print(venue_name_final02)
        # print(type(venue_name_final01))

        venue_country = str(venue_name_final.split()[-1]).strip("()")
        # print(venue_country)

        venue_total_sum = event_money_int
        # print(venue_total_sum)

        jockey_names = WebDriverWait(driver, 30).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "name")))

        # pro_quote01 = '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[
        # '2]/bf-marketview-runners-list[' \ '2]/div/div/div/table/tbody/tr[' pro_quote02 = ']/td[4]/button/div/span[1]'

        against_quote01 = '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[' \
                          '2]/bf-marketview-runners-list[' \
                          '2]/div/div/div/table/tbody/tr['
        against_quote02 = ']/td[5]/button/div/span[1]'

        quotes_final_against = []

        try:
            for jockey_place in range(len(jockey_names)):
                jokey = jockey_names[jockey_place]

                # pro_q = pro_quote01 + str(jockey_place + 1) + pro_quote02
                against_q = against_quote01 + str(jockey_place + 1) + against_quote02

                # quotes = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.XPATH, pro_q)))
                quotes01 = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.XPATH, against_q)))

                # print(jokey.text)
                # print(quotes.text)
                quotes_final_against.append(float(quotes01.text))
                # print(quotes01.text)

        except:
            'TimeoutException'

        # print(quotes_final_against)
        favorite_index = min(quotes_final_against)
        # print(favorite_index)

        favorite_jokey_name = jockey_names_list[quotes_final_against.index(favorite_index)]
        # print(favorite_jokey_name)

        # Calling the Test function to write the event in Excel
        test(venue_date_final01, venue_time_final, venue_name_final02, venue_country,
             venue_total_sum, favorite_index, favorite_jokey_name, race_to_check)

    else:
        pass


# The function that verifies if the race is writen in Excel, if it isn't it writes it in
# Also here Should be the Betting button
def test(venue_date_final01, venue_time_final, venue_name_final02, venue_country,
         venue_total_sum, favorite_index, favorite_jokey_name, race_to_check):

    # print(venue_date_final01, venue_time_final, venue_name_final02, venue_country,
    #       venue_total_sum, favorite_index, favorite_jokey_name)

    # Load the Excel file and make it active
    workbook_name = 'data.xlsx'
    wb = load_workbook(workbook_name)
    ws = wb.active

    # Because openpyxl doesn't have a module that let you select last n rows
    # I have to get the length of the C column for which I subtract the rows
    # that have to be checked, in this case I got the last 7 rows
    col_c = ws['c']
    len_table = int(len(col_c))
    # print print(len_table - 5)
    fifth_row = int(len_table - 7)

    # I bring the values form the race_to_start function in a list so we can write them in Excel
    race_final_data = [venue_date_final01, venue_time_final, venue_name_final02, venue_country,
                       venue_total_sum, favorite_index, favorite_jokey_name, race_to_check]

    # Select the columns and rows to be checked
    row = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=2, max_col=3)

    # Create an empty list of the last rows from Excel to be able to compare the item
    excel_last_rows = []

    for a, b in row:
        excel_last_rows.append([a.value, b.value])

    # The button that will place the bet will be in this condition as well
    if [venue_time_final, venue_name_final02] not in excel_last_rows:
        # print(venue_time_final, venue_name_final02)
        ws.append(race_final_data)

    wb.save(filename=workbook_name)


try:
    for race in race_link:
        driver.get(race)
        race_status = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "market-status-label")))
        if race_status.text == "Intră în desfăşurare":
            race_to_start()
        # elif race_status.text == 'Închis':
        #     race_finished()
        elif race_status.text == "În desfăşurare":
            pass
    race_finished()

except:
    'TimeoutException'


driver.quit()
