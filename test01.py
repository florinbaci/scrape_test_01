from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import ast
import openpyxl
import datetime
import re

# from openpyxl import Workbook
# from openpyxl import load_workbook
# from openpyxl.worksheet.table import Table, TableStyleInfo

# Setting up the environment and getting the web page raddy to scrap
driver = webdriver.Chrome('C:/Users/User/AppData/Local/Programs/Python/Python38/Scripts/chromedriver.exe')
driver.maximize_window()
url = "https://www.betfair.ro/exchange/plus/horse-racing/market/1.193538178?nodeId=31182619.1751"
driver.get(url)

# print(driver.page_source)

jockey_names = WebDriverWait(driver, 30).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "name")))
print(jockey_names)

driver.quit()
# race_to_check = []
# favorite = []

#
# def race_finished():
#     jockey_names = WebDriverWait(driver, 30).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "name")))
#     jockey_names_list = []
#     for jockey in jockey_names:
#         jokey_01 = jockey.text
#         jockey_names_list.append(jokey_01)
#         race_winner = WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.CLASS_NAME, "runner-winner")))
#         race_win01: str = '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[' \
#                           '1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[' \
#                           '2]/div/div/div/table/tbody/tr['
#         race_win02: str = ']/td/div[1]/div[2]/div'
#     # print(jockey_names_list)
#     position_final = []
#
#     for position in range(len(jockey_names)):
#         race_w = race_win01 + str(position + 1) + race_win02
#         try:
#             winner = WebDriverWait(driver, 5).until(ec.presence_of_element_located((By.XPATH, race_w)))
#             if winner.text == "Câştigător":
#                 position_final.append(position)
#             break
#         except:
#             "TimeoutException"
#     # This is what is needed at position 1
#     venue_name = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "venue-name")))
#     venue_name_final = venue_name.text
#     venue_name_final_race_fin01 = str(venue_name_final.split()[1:-1])
#     venue_name_final_race_fin02 = (' '.join(ast.literal_eval(venue_name_final_race_fin01)))
    # print(venue_name_final_race_fin02)
#
#     # This is what is needed at position 2
#     winner_final = jockey_names_list[int(position_final[0])]
#     winner_final_01 = re.sub('\d*\.*\n', '', winner_final)
#     print(winner_final_01)
#
#     venue_date_race_fin = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "event-date")))
#     venue_date_final_race_fin = venue_date_race_fin.text
#     venue_date_final_race_fin01 = venue_date_final_race_fin[3:]
#     # print(venue_date_final_race_fin01)
#
#     # This is what is needed at position 0
#     venue_time_race_fin = venue_name_final.split()[0]
#     venue_time_final_race_fin = datetime.datetime.strptime(venue_time_race_fin, '%H:%M').time()
#     # print(venue_time_final_race_fin)
#
#     race_final_data = [venue_time_final_race_fin, venue_name_final_race_fin02, winner_final]
#     race_to_check.append(race_final_data)
#     # print(race_final_data)
#     # favorite.append(winner_final)
#     # print(winner_final)
#     # print("in")
#
#     # driver.quit()
#
#
# # print(race_to_check)
# # print(favorite)
#
# # ref_workbook = openpyxl.load_workbook('data.xlsx')
# path = 'C:/Users/User/PycharmProjects/scrape_test_01/venv/data.xlsx'
# wb = openpyxl.load_workbook(path)
#
# workbook_name = 'data.xlsx'
# # wb = load_workbook(workbook_name)
# ws = wb.active
# # ws.title = "Races"
#
# # Take the last 5 rows of the Excel
# col_c = ws['c']
# len_table = int(len(col_c))
# # print(len_table)
# fifth_row = int(len_table - 5 + 1)
# # print(fifth_row)
#
# row = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=8, max_col=8)
#
# # Create an empty list to store the links that has to be checked
# links_to_check = []
#
# for link in row:
#     for link_text in link:
#         links_to_check.append(link_text.value)
#
# # print(links_to_check)
#
# # Create a list of lists with the last 5 events
# list_of_events = []
# list_of_dates = []
# list_of_races = []
# list_of_jockeys = []
#
# row_date = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=2, max_col=2)
# for race_date in row_date:
#     for event_date in race_date:
#         list_of_dates.append(event_date.value)
# # print(list_of_dates)
#
# row_races = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=3, max_col=3)
# for event_race in row_races:
#     for race_name in event_race:
#         list_of_races.append(race_name.value)
# # print(list_of_races)
#
# row_jockeys = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=7, max_col=7)
# for event_jokey in row_jockeys:
#     for race_jokey in event_jokey:
#         list_of_jockeys.append(race_jokey.value)
# # print(list_of_jockeys)
#
# for elem in range(len(list_of_dates)):
#     group_lists = [list_of_dates[elem], list_of_races[elem], list_of_jockeys[elem]]
#     list_of_events.append(group_lists)
# # print(list_of_events)
#
# # This two variables have to be scraped and checked
# # And then append the result to Excel
# # race_to_check = [datetime.time(7, 0), 'Benalla', 'The Nephew\nBilly Egan']
# # favorite = '1. The Nephew\nBilly Egan'
#
# # race_in_list = []
# #
# # for r in list_of_events:
# #     if race_to_check in list_of_events:
# #         index_found = list_of_events.index(race_to_check)
# #         print(index_found)
# #         race_in_list.append(race_to_check)
# #         break
# # # print(race_in_list[0][2])
# #
# # if favorite in race_in_list[0][2]:
# #     # Append 0 to Excel for a lost
# #     print("lost")
# # else:
# #     # Append 1 to Excel for a win
# #     print("won")
#
# driver = webdriver.Chrome('C:/Users/User/AppData/Local/Programs/Python/Python38/Scripts/chromedriver.exe')
# driver.maximize_window()
#
# try:
#     for race_fin in links_to_check:
#         driver.get(race_fin)
#         race_status = WebDriverWait(driver, 30).until(ec.presence_of_element_located((By.CLASS_NAME, "market-status-label")))
#         if race_status.text == "Intră în desfăşurare":
#             pass
#         elif race_status.text == 'Închis':
#             race_finished()
#         elif race_status.text == "În desfăşurare":
#             pass
#
#
# except:
#     'TimeoutException'
#
# driver.quit()
#
# # row_range = ws[fifth_row: len_table]
# # print(row_range)
#
# # col_b = int(fifth_row)
# # col_c = int(len_table)
# # col_bf = str('B') + str(col_b)
# # col_cf = str('C') + str(col_c)
# # # print(col_bf)
# # # print(col_cf)
#
# # time_race_list = []
# #
# # row_time_race = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=2, max_col=3)
# #
# # for time, race in row_time_race:
# #     time_race_list.append([time.value, race.value])
# #
# # jokey_list = []
# #
# # row_jokey = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=7, max_col=7)
# #
# # for jokey in row_jokey:
# #     for jokey_value in jokey:
# #         jokey_list.append(jokey_value.value)
# #
# # for time_race_jokey in range(len(time_race_list)):
# #     time_race_list[time_race_jokey].append(jokey_list[time_race_jokey])
# #
# # print(time_race_list)
# # print(jokey_list)
#
# # def index_in_list_of_lists(time_race_list, jokey_list):
# #     for i, lst in enumerate(time_race_list):
# #         if jokey_list in lst:
# #             break
# #         else:
# #             raise ValueError("%s not in list_of_lists" %jokey_list)
# #
# #     return i, lst.index(jokey_list)
# #
# #
# # print(index_in_list_of_lists(time_race_list, jokey_list))
# #
# # h = index_in_list_of_lists(time_race_list, jokey_list)
# # print(h[0])
#
# # if [datetime.time(7, 10), 'Newcastle', 'Jesse James\nBrodie Loy'] in time_race_list:
# #     print(time_race_list.index([datetime.time(7, 10), 'Newcastle', 'Jesse James\nBrodie Loy']))
# #     print('found')
# # else:
# #     print('not found')
# #     race_to_check.append([a.value, b.value, c.value, e.value, f.value])
# # for a, b, c, d, e, f in row_01:
# #     race_to_check.append([a.value, b.value, c.value, e.value, f.value])
# # if [datetime.time(6, 50), 'Queanbeyan', 'AUS', 'Milamoo\nAmy Mc Lucas'] in race_to_check:
# #     print("found")
# #     break
# # print(a.value, b.value)
# # if "Philip" in a.value and "Gent" in b.value:
# #     print(a.value, b.value)
# #     # ws.cell(row=1, column=9).value = 23
# #     # print(b)
# #     row_no = str(b)
# #     row_final = int(row_no[16:-1])
# #     # ws.cell(row=row_final, column=9).value = 99
# #
# #     pattern = int(re.findall(r'\d+', row_no)[-1])
# #     print(pattern)
# #     ws.cell(row=pattern, column=9).value = 89
#
# # row_02 = ws.iter_rows(min_row=fifth_row, max_row=len_table, min_col=7,  max_col=7)
# # #
# # for c in row_02:
# #     for jokey in c:
# #         print(race_to_check[0])
# # race_to_check.insert(0, [jokey.value])
# # print(jokey.value)
#
# # race_to_check = [[] for _ in range(7)]
# # race_to_check[0:-1].append([jokey.value])
#
# # print(race_to_check)
#
# # wb.save(filename=workbook_name)
